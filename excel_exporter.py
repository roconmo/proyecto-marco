"""
excel_exporter.py
Lógica de escritura del output en la plantilla Excel del usuario.

Reglas de escritura:
- Trabajar siempre sobre la hoja "Hoja1" del Excel proporcionado.
- Leer la cabecera existente y localizar columnas por nombre (no por índice fijo).
- ANTES de escribir cada línea, buscar el código MAVY en la columna "Modelo":
    · Si NO existe → escribir "N" en "Cód. concepto" y "Nº".
    · Si SÍ existe → escribir el código en "Cód. concepto" y "Nº".
- Colorear cada fila nueva en verde claro (C6EFCE).
- Nunca crear hojas nuevas, nunca duplicar cabecera, nunca sobrescribir datos.
"""

import io
from typing import Any, Dict, List, Set, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

# Hoja de destino fija
HOJA_DESTINO = "Hoja1"

# Columnas que provienen directamente del parser
COLUMNAS_PARSER: List[str] = [
    "Descripción",
    "Marca",
    "Modelo",
    "Medidas",
    "Cantidad",
    "Coste vigente",
    "Coste unitario (DL)",
    "Precio venta excl. IVA",
    "% Descuento línea",
    "Margen",
    "Importe",
    "Importe línea excl. IVA",
]

# Columnas calculadas por el exportador (NO provienen del parser)
# Su valor se determina mediante la búsqueda del código MAVY en "Modelo"
COLUMNAS_LOOKUP: List[str] = ["Cód. concepto", "Nº"]

# Todas las columnas que el exportador puede necesitar escribir
COLUMNAS_OBJETIVO: List[str] = COLUMNAS_PARSER + COLUMNAS_LOOKUP

# Relleno verde claro para filas nuevas
_FILL_VERDE = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid",
)

# Relleno y fuente para filas con advertencia (marca/código de lectura dudosa)
_FILL_ROJO = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid",
)
_FONT_ROJO_NEGRITA = Font(bold=True, color="9C0006")


# ---------------------------------------------------------------------------
# Función pública principal
# ---------------------------------------------------------------------------

def exportar_a_excel(
    filas: List[Dict[str, Any]],
    excel_bytes: bytes,
    hoja: str = HOJA_DESTINO,
) -> Tuple[bytes, List[str]]:
    """
    Escribe las filas del parser en "Hoja1" aplicando la lógica especial de MAVY.

    Flujo por cada fila:
      1. Leer el código MAVY desde fila["Modelo"].
      2. Buscar ese código en los valores existentes de la columna "Modelo".
         · NO encontrado → "Cód. concepto" = "N"  /  "Nº" = "N"
         · SÍ encontrado → "Cód. concepto" = código  /  "Nº" = código
      3. Escribir los datos en las columnas correspondientes.
      4. Colorear toda la fila nueva en verde claro (C6EFCE).

    Args:
        filas:       Lista de dicts devuelta por el parser.
        excel_bytes: Contenido del Excel original en bytes.
        hoja:        Nombre de la hoja destino (por defecto "Hoja1").

    Returns:
        Tuple (excel_bytes_actualizado, columnas_no_encontradas)
        · excel_bytes_actualizado : bytes listos para st.download_button.
        · columnas_no_encontradas : columnas objetivo ausentes en la cabecera.
    """
    if not filas:
        return excel_bytes, []

    wb = load_workbook(io.BytesIO(excel_bytes))

    # Validar que la hoja existe
    if hoja not in wb.sheetnames:
        raise ValueError(
            f"La hoja '{hoja}' no existe en el Excel seleccionado. "
            f"Hojas disponibles: {wb.sheetnames}"
        )

    ws = wb[hoja]

    # -----------------------------------------------------------------------
    # 1. Leer cabecera → mapa {nombre_columna: índice_1based}
    # -----------------------------------------------------------------------
    mapa_cabecera = _leer_cabecera(ws)

    # -----------------------------------------------------------------------
    # 2. Mapear columnas del parser
    # -----------------------------------------------------------------------
    mapa_parser, faltantes_parser = _mapear_lista(mapa_cabecera, COLUMNAS_PARSER)

    # -----------------------------------------------------------------------
    # 3. Mapear columnas de lookup (Cód. concepto, Nº)
    # -----------------------------------------------------------------------
    mapa_lookup, faltantes_lookup = _mapear_lista(mapa_cabecera, COLUMNAS_LOOKUP)

    # Acumular todas las faltantes para informar al usuario
    columnas_faltantes: List[str] = faltantes_parser + faltantes_lookup

    # Si no hay ninguna columna escribible, no tiene sentido continuar
    if not mapa_parser and not mapa_lookup:
        raise ValueError(
            f"Ninguna columna objetivo encontrada en '{hoja}'. "
            "Comprueba que la plantilla es correcta."
        )

    # -----------------------------------------------------------------------
    # 4. Leer valores existentes en "Modelo" ANTES de escribir
    #    (para el lookup; se lee una sola vez para todo el lote)
    # -----------------------------------------------------------------------
    col_idx_modelo = mapa_cabecera.get("Modelo")
    modelos_existentes: Set[str] = (
        _leer_valores_columna(ws, col_idx_modelo)
        if col_idx_modelo
        else set()
    )

    # -----------------------------------------------------------------------
    # 5. Detectar primera fila libre
    # -----------------------------------------------------------------------
    ultima_fila = find_last_row(ws)
    fila_inicio = ultima_fila + 1

    # Número de columnas del Excel (para colorear toda la fila)
    num_cols_excel = max(mapa_cabecera.values()) if mapa_cabecera else 1

    # -----------------------------------------------------------------------
    # 6. Escribir filas
    # -----------------------------------------------------------------------
    for offset, fila_dict in enumerate(filas):
        fila_destino = fila_inicio + offset

        # — Lógica de lookup —
        codigo_mavy = str(fila_dict.get("Modelo", "")).strip()
        valor_lookup = (
            codigo_mavy
            if (codigo_mavy and codigo_mavy in modelos_existentes)
            else "N"
        )

        # — Escribir columnas del parser —
        for nombre_col, col_idx in mapa_parser.items():
            valor = fila_dict.get(nombre_col, None)
            # Convertir "" a None para no escribir cadenas vacías en la celda
            ws.cell(
                row=fila_destino,
                column=col_idx,
                value=valor if valor != "" else None,
            )

        # — Escribir columnas de lookup (Cód. concepto / Nº) —
        for nombre_col, col_idx in mapa_lookup.items():
            ws.cell(row=fila_destino, column=col_idx, value=valor_lookup)

        # — Colorear toda la fila en verde —
        for col_idx in range(1, num_cols_excel + 1):
            ws.cell(row=fila_destino, column=col_idx).fill = _FILL_VERDE

        # — Si hay advertencia, pintar solo Marca y Modelo en rojo negrita —
        if bool(fila_dict.get("_advertencia", False)):
            for nombre_col in ("Marca", "Modelo"):
                col_idx = mapa_parser.get(nombre_col)
                if col_idx:
                    cell = ws.cell(row=fila_destino, column=col_idx)
                    cell.fill = _FILL_ROJO
                    cell.font = _FONT_ROJO_NEGRITA

    # -----------------------------------------------------------------------
    # 7. Serializar y devolver
    # -----------------------------------------------------------------------
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read(), columnas_faltantes


# ---------------------------------------------------------------------------
# Funciones auxiliares
# ---------------------------------------------------------------------------

def find_last_row(ws) -> int:
    """
    Devuelve el número de la última fila con algún valor real.
    Devuelve 0 si la hoja está completamente vacía.

    Más fiable que ws.max_row cuando hay celdas con formato pero sin valor
    (p.ej. rangos estilizados vacíos en plantillas Excel).
    """
    for row in reversed(range(1, ws.max_row + 1)):
        if any(cell.value for cell in ws[row]):
            return row
    return 0


def _leer_cabecera(ws) -> Dict[str, int]:
    """
    Lee la fila 1 y devuelve {nombre_columna_strip: índice_columna_1based}.
    Ignora celdas vacías.
    """
    mapa: Dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            mapa[str(cell.value).strip()] = cell.column
    return mapa


def _mapear_lista(
    mapa_cabecera: Dict[str, int],
    columnas: List[str],
) -> Tuple[Dict[str, int], List[str]]:
    """
    Cruza una lista de columnas objetivo con el mapa de cabecera real.

    Returns:
        (encontradas, faltantes)
        · encontradas : {nombre: índice} para las que sí existen en la cabecera
        · faltantes   : nombres que NO se encontraron
    """
    encontradas: Dict[str, int] = {}
    faltantes: List[str] = []
    for col in columnas:
        if col in mapa_cabecera:
            encontradas[col] = mapa_cabecera[col]
        else:
            faltantes.append(col)
    return encontradas, faltantes


def _leer_valores_columna(ws, col_idx: int) -> Set[str]:
    """
    Lee todos los valores no nulos de una columna (excepto fila 1 = cabecera).
    Devuelve un set de strings con strip() para comparaciones limpias.

    Args:
        ws:      Worksheet de openpyxl
        col_idx: Índice de columna 1-based

    Returns:
        Set de strings normalizados
    """
    valores: Set[str] = set()
    ultima = find_last_row(ws)
    for row in range(2, ultima + 1):  # Fila 2 en adelante (salta cabecera)
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            valores.add(str(cell.value).strip())
    return valores


def inspeccionar_cabecera(excel_bytes: bytes, hoja: str = HOJA_DESTINO) -> List[str]:
    """
    Utilidad de apoyo: devuelve la lista de nombres de columnas de la fila 1.
    Usada en la UI para mostrar qué columnas tiene realmente la plantilla.

    Args:
        excel_bytes: Contenido del Excel en bytes.
        hoja:        Nombre de la hoja a inspeccionar.

    Returns:
        Lista de strings con los nombres de columnas (en orden).
    """
    wb = load_workbook(io.BytesIO(excel_bytes))
    if hoja not in wb.sheetnames:
        return []
    ws = wb[hoja]
    return [
        str(cell.value).strip()
        for cell in ws[1]
        if cell.value is not None
    ]
