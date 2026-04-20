"""
verificar.py
Comprobación final: contrasta los datos extraídos del PDF con las filas
recién escritas en el Excel para detectar discrepancias.

La comparación se hace fila a fila sobre los campos numéricos clave
(Cantidad, Coste vigente, Importe) y los campos de identificación
(Marca, Modelo, Descripción). Las diferencias se devuelven como una
lista de dicts para mostrar en la UI.
"""

from __future__ import annotations

import io
import math
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

from excel_exporter import find_last_row, HOJA_DESTINO

# Campos que se comparan, con tolerancia para floats
_CAMPOS_ID = ["Marca", "Modelo", "Descripción"]
_CAMPOS_NUM = ["Cantidad", "Coste vigente", "Importe"]
_TOLERANCIA = 0.02  # diferencia máxima aceptable en campos numéricos


def verificar(
    filas_pdf: List[Dict[str, Any]],
    excel_bytes_resultado: bytes,
    n_filas_escritas: int,
    hoja: str = HOJA_DESTINO,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Compara las `n_filas_escritas` últimas filas del Excel con `filas_pdf`.

    Returns:
        (discrepancias, resumen)
        · discrepancias : lista de dicts con detalles de cada diferencia
        · resumen       : lista de strings para mostrar como informe
    """
    discrepancias: List[Dict[str, Any]] = []

    # --- Leer las filas recién escritas del Excel ---
    wb = load_workbook(io.BytesIO(excel_bytes_resultado))
    ws = wb[hoja]

    # Cabecera
    cabecera = {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }

    ultima = find_last_row(ws)
    primera_nueva = ultima - n_filas_escritas + 1

    filas_excel: List[Dict[str, Any]] = []
    for row_idx in range(primera_nueva, ultima + 1):
        fila = {}
        for nombre_col, col_idx in cabecera.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            fila[nombre_col] = val
        filas_excel.append(fila)

    # --- Comprobación de conteo ---
    if len(filas_excel) != len(filas_pdf):
        discrepancias.append({
            "tipo": "conteo",
            "fila": "-",
            "campo": "Número de filas",
            "pdf": len(filas_pdf),
            "excel": len(filas_excel),
        })

    # --- Comparación campo a campo ---
    n = min(len(filas_pdf), len(filas_excel))
    for i in range(n):
        fp = filas_pdf[i]
        fe = filas_excel[i]
        fila_num = i + 1

        for campo in _CAMPOS_ID:
            val_pdf = str(fp.get(campo, "") or "").strip()
            val_excel = str(fe.get(campo, "") or "").strip()
            if val_pdf != val_excel:
                discrepancias.append({
                    "tipo": "texto",
                    "fila": fila_num,
                    "campo": campo,
                    "pdf": val_pdf,
                    "excel": val_excel,
                })

        for campo in _CAMPOS_NUM:
            v_pdf = _to_float(fp.get(campo, ""))
            v_excel = _to_float(fe.get(campo, ""))
            if v_pdf is None and v_excel is None:
                continue
            if v_pdf is None or v_excel is None or abs(v_pdf - v_excel) > _TOLERANCIA:
                discrepancias.append({
                    "tipo": "número",
                    "fila": fila_num,
                    "campo": campo,
                    "pdf": v_pdf,
                    "excel": v_excel,
                })

    # --- Construir resumen legible ---
    resumen: List[str] = []
    if not discrepancias:
        resumen.append(f"✔ Verificación correcta — {n} filas comparadas sin diferencias.")
    else:
        resumen.append(
            f"⚠ Se encontraron {len(discrepancias)} discrepancia(s) en {n} filas comparadas."
        )

    return discrepancias, resumen


def _to_float(valor: Any) -> float | None:
    if valor is None or str(valor).strip() == "":
        return None
    # Si ya es numérico, no aplicar sustitución de separadores europeos
    if isinstance(valor, (int, float)):
        return None if math.isnan(float(valor)) else float(valor)
    try:
        s = str(valor).strip()
        # Formato europeo: punto = separador de miles, coma = decimal
        v = float(s.replace(".", "").replace(",", "."))
        return None if math.isnan(v) else v
    except (ValueError, TypeError):
        return None


def discrepancias_a_dataframe(discrepancias: List[Dict[str, Any]]) -> "pd.DataFrame":
    if not discrepancias:
        return pd.DataFrame()
    return pd.DataFrame(discrepancias)[["fila", "campo", "pdf", "excel", "tipo"]]
